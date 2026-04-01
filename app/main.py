from flask import Flask, jsonify, request, send_file, session, send_from_directory
from io import BytesIO
from openpyxl import Workbook, load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
import os
import json
import unicodedata

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
STATIC_DIR = os.path.join(BASE_DIR, "static")
DATA_DIR = os.path.join(BASE_DIR, "data")
DB_PATH = os.path.join(DATA_DIR, "db.json")
EXCEL_PATH = os.path.join(DATA_DIR, "base_sage.xlsx")

app = Flask(__name__, static_folder=STATIC_DIR, static_url_path="")
app.secret_key = os.environ.get("SECRET_KEY", "grupobolhao-secret")
APP_PIN = os.environ.get("APP_PIN", "1234")

DEFAULT_DB = {
    "purchases": [],
    "cash_state": {
        "talho": {
            "date": "",
            "start": 0,
            "inCash": 0,
            "inMb": 0,
            "inMbway": 0,
            "inOther": 0,
            "out": 0,
            "obs": "",
            "notes": {"500": 0, "200": 0, "100": 0, "50": 0, "20": 0, "10": 0, "5": 0},
            "coins": {"2": 0, "1": 0, "0.5": 0, "0.2": 0, "0.1": 0, "0.05": 0, "0.02": 0, "0.01": 0}
        },
        "cong": {
            "date": "",
            "start": 0,
            "inCash": 0,
            "inMb": 0,
            "inMbway": 0,
            "inOther": 0,
            "out": 0,
            "obs": "",
            "notes": {"500": 0, "200": 0, "100": 0, "50": 0, "20": 0, "10": 0, "5": 0},
            "coins": {"2": 0, "1": 0, "0.5": 0, "0.2": 0, "0.1": 0, "0.05": 0, "0.02": 0, "0.01": 0}
        }
    },
    "next_purchase_id": 1
}


def ensure_data():
    os.makedirs(DATA_DIR, exist_ok=True)
    if not os.path.exists(DB_PATH):
        save_db(DEFAULT_DB)


def save_db(db):
    os.makedirs(DATA_DIR, exist_ok=True)
    with open(DB_PATH, "w", encoding="utf-8") as f:
        json.dump(db, f, ensure_ascii=False, indent=2)


def load_db():
    ensure_data()

    if not os.path.exists(DB_PATH):
        save_db(DEFAULT_DB)
        return json.loads(json.dumps(DEFAULT_DB))

    try:
        with open(DB_PATH, "r", encoding="utf-8") as f:
            db = json.load(f)
    except Exception:
        save_db(DEFAULT_DB)
        return json.loads(json.dumps(DEFAULT_DB))

    changed = False

    if "purchases" not in db or not isinstance(db["purchases"], list):
        db["purchases"] = []
        changed = True

    if "cash_state" not in db or not isinstance(db["cash_state"], dict):
        db["cash_state"] = DEFAULT_DB["cash_state"]
        changed = True

    if "next_purchase_id" not in db or not isinstance(db["next_purchase_id"], int):
        existing_ids = [item.get("id", 0) for item in db.get("purchases", []) if isinstance(item, dict)]
        db["next_purchase_id"] = (max(existing_ids) + 1) if existing_ids else 1
        changed = True

    if changed:
        save_db(db)

    return db


def require_login():
    if not session.get("ok"):
        return jsonify({"error": "Não autenticado"}), 401
    return None


def parse_amount(value):
    if value is None or value == "":
        return 0.0
    return round(float(str(value).replace(",", ".")), 2)


def normalize_text(value):
    text = str(value or "").strip().lower()
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode("ascii")
    return text


def find_sheet_name(workbook, expected_name):
    expected = normalize_text(expected_name)
    for sheet_name in workbook.sheetnames:
        if normalize_text(sheet_name) == expected:
            return sheet_name
    return None


def load_excel_master():
    if not os.path.exists(EXCEL_PATH):
        return {"articles": [], "suppliers": []}

    wb = load_workbook(EXCEL_PATH, data_only=True)
    articles = []
    suppliers = []

    artigos_sheet = find_sheet_name(wb, "artigos")
    fornecedores_sheet = find_sheet_name(wb, "fornecedores")

    if artigos_sheet:
        ws = wb[artigos_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [normalize_text(h) for h in rows[0]]
            for row in rows[1:]:
                item = dict(zip(headers, row))
                code = str(item.get("codigo", "") or "").strip()
                name = str(item.get("nome", "") or "").strip()
                if code and name:
                    articles.append({"code": code, "name": name})

    if fornecedores_sheet:
        ws = wb[fornecedores_sheet]
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            headers = [normalize_text(h) for h in rows[0]]
            for row in rows[1:]:
                item = dict(zip(headers, row))
                code = str(item.get("codigo", "") or "").strip()
                name = str(item.get("nome", "") or "").strip()
                if code and name:
                    suppliers.append({"code": code, "name": name})

    return {"articles": articles, "suppliers": suppliers}


def purchase_state(item):
    qty_to_buy = parse_amount(item.get("qty_to_buy", 0))
    qty_bought = parse_amount(item.get("qty_bought", 0))

    if qty_bought <= 0:
        return "Por comprar"
    if qty_bought < qty_to_buy:
        return "Parcial"
    return "Comprado"


def calc_cash_summary(section_data):
    notes_total = 0.0
    coins_total = 0.0

    note_values = {"500": 500, "200": 200, "100": 100, "50": 50, "20": 20, "10": 10, "5": 5}
    coin_values = {"2": 2, "1": 1, "0.5": 0.5, "0.2": 0.2, "0.1": 0.1, "0.05": 0.05, "0.02": 0.02, "0.01": 0.01}

    for key, value in note_values.items():
        notes_total += parse_amount(section_data.get("notes", {}).get(key, 0)) * value

    for key, value in coin_values.items():
        coins_total += parse_amount(section_data.get("coins", {}).get(key, 0)) * value

    real = round(notes_total + coins_total, 2)
    expected = round(
        parse_amount(section_data.get("start", 0))
        + parse_amount(section_data.get("inCash", 0))
        + parse_amount(section_data.get("inMb", 0))
        + parse_amount(section_data.get("inMbway", 0))
        + parse_amount(section_data.get("inOther", 0))
        - parse_amount(section_data.get("out", 0)),
        2
    )
    diff = round(real - expected, 2)

    status = "Certo"
    if abs(diff) >= 0.01:
        status = "Sobra em caixa" if diff > 0 else "Falta em caixa"

    return {
        "notes_total": notes_total,
        "coins_total": coins_total,
        "real": real,
        "expected": expected,
        "diff": diff,
        "status": status
    }


@app.route("/")
def home():
    return send_from_directory(STATIC_DIR, "index.html")


@app.route("/<path:path>")
def static_files(path):
    full = os.path.join(STATIC_DIR, path)
    if os.path.exists(full):
        return send_from_directory(STATIC_DIR, path)
    return send_from_directory(STATIC_DIR, "index.html")


@app.route("/api/login", methods=["POST"])
def api_login():
    data = request.get_json(silent=True) or {}
    pin = str(data.get("pin", "")).strip()

    if pin == APP_PIN:
        session["ok"] = True
        return jsonify({"ok": True})

    return jsonify({"ok": False, "error": "PIN inválido"}), 401


@app.route("/api/status")
def api_status():
    return jsonify({"logged_in": bool(session.get("ok"))})


@app.route("/api/logout", methods=["POST"])
def api_logout():
    session.clear()
    return jsonify({"ok": True})


@app.route("/api/master/articles")
def api_master_articles():
    auth = require_login()
    if auth:
        return auth
    return jsonify(load_excel_master()["articles"])


@app.route("/api/master/suppliers")
def api_master_suppliers():
    auth = require_login()
    if auth:
        return auth
    return jsonify(load_excel_master()["suppliers"])


@app.route("/api/purchases", methods=["GET", "POST"])
def api_purchases():
    auth = require_login()
    if auth:
        return auth

    db = load_db()

    if request.method == "GET":
        return jsonify(db.get("purchases", []))

    data = request.get_json(silent=True) or {}

    item = {
        "id": db.get("next_purchase_id", 1),
        "code": str(data.get("code", "")).strip(),
        "name": str(data.get("name", "")).strip(),
        "supplier_code": str(data.get("supplier_code", "")).strip(),
        "supplier": str(data.get("supplier", "")).strip(),
        "qty_to_buy": parse_amount(data.get("qty_to_buy", 0)),
        "qty_bought": parse_amount(data.get("qty_bought", 0)),
        "unit": str(data.get("unit", "")).strip(),
        "priority": str(data.get("priority", "")).strip()
    }

    db["purchases"].append(item)
    db["next_purchase_id"] = db.get("next_purchase_id", 1) + 1
    save_db(db)

    return jsonify({"ok": True, "item": item})


@app.route("/api/purchases/<int:item_id>/complete", methods=["POST"])
def api_purchase_complete(item_id):
    auth = require_login()
    if auth:
        return auth

    db = load_db()

    for item in db["purchases"]:
        if item["id"] == item_id:
            item["qty_bought"] = item["qty_to_buy"]
            save_db(db)
            return jsonify({"ok": True, "item": item})

    return jsonify({"error": "Artigo não encontrado"}), 404


@app.route("/api/purchases/<int:item_id>", methods=["DELETE"])
def api_purchase_delete(item_id):
    auth = require_login()
    if auth:
        return auth

    db = load_db()
    before = len(db["purchases"])
    db["purchases"] = [p for p in db["purchases"] if p["id"] != item_id]

    if len(db["purchases"]) == before:
        return jsonify({"error": "Artigo não encontrado"}), 404

    save_db(db)
    return jsonify({"ok": True})


@app.route("/api/cash-state", methods=["GET", "POST"])
def api_cash_state():
    auth = require_login()
    if auth:
        return auth

    db = load_db()

    if request.method == "GET":
        return jsonify(db["cash_state"])

    data = request.get_json(silent=True) or {}
    db["cash_state"] = data
    save_db(db)

    return jsonify({"ok": True})


@app.route("/api/export/purchases/excel")
def export_purchases_excel():
    auth = require_login()
    if auth:
        return auth

    db = load_db()
    purchases = db["purchases"]

    wb = Workbook()
    ws = wb.active
    ws.title = "compras"

    ws.append([
        "ID", "Código Artigo", "Artigo", "Código Fornecedor", "Fornecedor",
        "Qtd Comprar", "Qtd Comprada", "Unidade", "Prioridade", "Estado"
    ])

    for p in purchases:
        ws.append([
            p.get("id"),
            p.get("code"),
            p.get("name"),
            p.get("supplier_code"),
            p.get("supplier"),
            p.get("qty_to_buy"),
            p.get("qty_bought"),
            p.get("unit"),
            p.get("priority"),
            purchase_state(p)
        ])

    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)

    return send_file(
        mem,
        as_attachment=True,
        download_name="compras.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/export/purchases/pdf")
def export_purchases_pdf():
    auth = require_login()
    if auth:
        return auth

    db = load_db()
    purchases = db["purchases"]

    mem = BytesIO()
    c = canvas.Canvas(mem, pagesize=A4)
    width, height = A4

    y = height - 40
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "ARTIGOS A COMPRAR")
    y -= 25

    c.setFont("Helvetica", 10)

    if not purchases:
        c.drawString(40, y, "Sem artigos registados.")
    else:
        for p in purchases:
            estado = purchase_state(p)
            linhas = [
                f"Código: {p.get('code', '')} | Artigo: {p.get('name', '')}",
                f"Fornecedor: {p.get('supplier_code', '')} - {p.get('supplier', '')}",
                f"Qtd comprar: {p.get('qty_to_buy', 0)} | Qtd comprada: {p.get('qty_bought', 0)} | Unidade: {p.get('unit', '')}",
                f"Prioridade: {p.get('priority', '')} | Estado: {estado}"
            ]

            for linha in linhas:
                c.drawString(40, y, linha)
                y -= 15
                if y < 50:
                    c.showPage()
                    y = height - 40
                    c.setFont("Helvetica", 10)

            y -= 10

    c.save()
    mem.seek(0)

    return send_file(
        mem,
        as_attachment=True,
        download_name="compras.pdf",
        mimetype="application/pdf"
    )


@app.route("/api/export/cash/excel")
def export_cash_excel():
    auth = require_login()
    if auth:
        return auth

    db = load_db()
    cash_state = db["cash_state"]

    wb = Workbook()
    ws = wb.active
    ws.title = "fecho_caixa"

    ws.append([
        "Secção", "Data", "Fundo Inicial", "Entradas Dinheiro", "Entradas Multibanco",
        "Entradas MBWay", "Outras Entradas", "Saídas",
        "Total Notas", "Total Moedas", "Caixa Real", "Saldo Teórico", "Diferença", "Estado", "Observações"
    ])

    for secao, label in [("talho", "Talho"), ("cong", "Congelados")]:
        section_data = cash_state.get(secao, {})
        summary = calc_cash_summary(section_data)

        ws.append([
            label,
            section_data.get("date", ""),
            section_data.get("start", 0),
            section_data.get("inCash", 0),
            section_data.get("inMb", 0),
            section_data.get("inMbway", 0),
            section_data.get("inOther", 0),
            section_data.get("out", 0),
            summary["notes_total"],
            summary["coins_total"],
            summary["real"],
            summary["expected"],
            summary["diff"],
            summary["status"],
            section_data.get("obs", "")
        ])

    mem = BytesIO()
    wb.save(mem)
    mem.seek(0)

    return send_file(
        mem,
        as_attachment=True,
        download_name="fecho_caixa.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


@app.route("/api/export/cash/pdf")
def export_cash_pdf():
    auth = require_login()
    if auth:
        return auth

    db = load_db()
    cash_state = db["cash_state"]

    mem = BytesIO()
    c = canvas.Canvas(mem, pagesize=A4)
    width, height = A4

    y = height - 40
    c.setFont("Helvetica-Bold", 14)
    c.drawString(40, y, "FECHO DE CAIXA")
    y -= 25

    c.setFont("Helvetica", 10)

    for secao, label in [("talho", "Talho"), ("cong", "Congelados")]:
        section_data = cash_state.get(secao, {})
        summary = calc_cash_summary(section_data)

        linhas = [
            f"Secção: {label}",
            f"Data: {section_data.get('date', '')}",
            f"Fundo inicial: {section_data.get('start', 0)} €",
            f"Entradas dinheiro: {section_data.get('inCash', 0)} €",
            f"Entradas multibanco: {section_data.get('inMb', 0)} €",
            f"Entradas MBWay: {section_data.get('inMbway', 0)} €",
            f"Outras entradas: {section_data.get('inOther', 0)} €",
            f"Saídas: {section_data.get('out', 0)} €",
            f"Total notas: {summary['notes_total']} €",
            f"Total moedas: {summary['coins_total']} €",
            f"Caixa real: {summary['real']} €",
            f"Saldo teórico: {summary['expected']} €",
            f"Diferença: {summary['diff']} €",
            f"Estado: {summary['status']}",
            f"Observações: {section_data.get('obs', '')}"
        ]

        for linha in linhas:
            c.drawString(40, y, linha)
            y -= 15
            if y < 50:
                c.showPage()
                y = height - 40
                c.setFont("Helvetica", 10)

        y -= 10

    c.save()
    mem.seek(0)

    return send_file(
        mem,
        as_attachment=True,
        download_name="fecho_caixa.pdf",
        mimetype="application/pdf"
    )


if __name__ == "__main__":
    ensure_data()
    app.run(host="0.0.0.0", port=5000)
